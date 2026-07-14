"""XLSX parser — openpyxl-backed, multi-table per sheet via empty-row boundary detection.

Empty rows act as separators between logical tables on the same sheet.
Single-row groups (titles, section labels) are counted toward word_count but
not emitted as tables.  Formulas: openpyxl data_only=True reads the cached
computed value; if the workbook was never opened in Excel the cache is empty
and formula cells arrive as None → converted to "".  A warning is surfaced in
raw_text_preview when the blank rate in a table exceeds 40%.
"""
from __future__ import annotations
from pathlib import Path

from models.events import ExtractedTable
from pipelines.base import StageResult

_XLSX_ROW_LIMIT = 1000  # per sheet — surface truncation rather than silently cut
_MAX_TABLE_ROWS = 25    # rows kept per table in preview / as_json


def _to_markdown(headers: list[str], rows: list[list[str]]) -> str:
    if not headers:
        return ""
    sep = ["---"] * len(headers)
    lines = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join(sep) + " |",
    ]
    for row in rows[:_MAX_TABLE_ROWS]:
        padded = list(row) + [""] * max(0, len(headers) - len(row))
        lines.append("| " + " | ".join(str(c) for c in padded[: len(headers)]) + " |")
    return "\n".join(lines)


def row_groups(ws) -> tuple[list[list[list[str]]], int]:
    """Split a worksheet into logical table groups separated by empty rows.

    Returns (groups, rows_read).  Each group is a list of non-empty rows.
    Caller receives the raw row_count so it can detect truncation.
    """
    groups: list[list[list[str]]] = []
    current: list[list[str]] = []
    row_count = 0

    for raw_row in ws.iter_rows(values_only=True):
        row_count += 1
        if row_count > _XLSX_ROW_LIMIT:
            break
        cells = [str(c).strip() if c is not None else "" for c in raw_row]
        while cells and not cells[-1]:
            cells.pop()

        if any(cells):
            current.append(cells)
        else:
            if current:
                groups.append(current)
                current = []

    if current:
        groups.append(current)

    return groups, row_count


def pad_rows(rows: list[list[str]], width: int) -> list[list[str]]:
    return [(r + [""] * width)[:width] for r in rows]


async def parse(filepath: Path, mime: str, result_fn) -> StageResult:
    """Parse an XLSX/XLS file and return a StageResult via result_fn.

    result_fn matches the signature of stage_03_parser._result so the caller
    can keep its own shared helper.
    """
    import openpyxl

    wb = openpyxl.load_workbook(str(filepath), data_only=True, read_only=True)
    sheet_names = wb.sheetnames
    tables: list[ExtractedTable] = []
    word_count = 0
    notes: list[str] = []

    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        groups, rows_read = row_groups(ws)

        if rows_read > _XLSX_ROW_LIMIT:
            notes.append(f"Sheet '{sheet_name}': truncated at {_XLSX_ROW_LIMIT} rows")

        sheet_tbl_idx = 0

        for group in groups:
            if len(group) < 2:
                for row in group:
                    word_count += sum(len(c.split()) for c in row if c)
                continue

            headers = group[0]
            body = pad_rows(group[1:], len(headers))

            total_cells = sum(len(r) for r in body)
            blank_cells = sum(1 for r in body for c in r if not c)
            blank_pct = blank_cells / total_cells if total_cells else 0
            if blank_pct > 0.4 and total_cells > 10:
                notes.append(
                    f"Sheet '{sheet_name}' table {sheet_tbl_idx + 1}: "
                    f"{blank_pct:.0%} blank cells — formula cache may be empty"
                )

            for row in group:
                word_count += sum(len(c.split()) for c in row if c)

            sheet_tbl_idx += 1
            tbl_id = (
                f"sheet_{sheet_name}_t{sheet_tbl_idx}"
                if sheet_tbl_idx > 1
                else f"sheet_{sheet_name}"
            )

            tables.append(ExtractedTable(
                id=tbl_id,
                page=None,
                headers=headers,
                rows=body,
                as_markdown=_to_markdown(headers, body),
                as_json=[dict(zip(headers, row)) for row in body[:_MAX_TABLE_ROWS]],
            ))

    wb.close()

    sheet_summary = f"{len(sheet_names)} sheet(s): {', '.join(sheet_names[:6])}"
    if notes:
        sheet_summary += " · " + " · ".join(notes)

    return result_fn("openpyxl", mime, [], tables, word_count, None, 0, sheet_summary)
